# ingest.py
import io
from typing import Dict, Any, Optional, Tuple, List
from datetime import date, timedelta
import re

import pandas as pd
from bs4 import BeautifulSoup
import requests
from docx import Document as DocxDocument
from beanie import PydanticObjectId

from pdfminer.high_level import extract_text as pdf_extract_text
from PyPDF2 import PdfReader
from openpyxl import load_workbook

from models import (
    Funcionario, Projeto, Tarefa,
    TarefaCreate, PrioridadeTarefa, StatusTarefa, CondicaoTarefa
)

# =========================
# Helpers gerais
# =========================
def _normalize_bool(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "t", "sim", "yes", "y", "concluida", "concluído", "done"}

def _fetch_url_text(u: str) -> str:
    """Busca texto essencial de uma URL (HTML)."""
    try:
        r = requests.get(u, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        main = soup.find("main") or soup.find("article") or soup.body
        return main.get_text(separator="\n", strip=True)[:15000] if main else soup.get_text(separator="\n", strip=True)[:15000]
    except Exception:
        return ""

def _fetch_bytes(u: str) -> Optional[bytes]:
    try:
        r = requests.get(u, timeout=40)
        r.raise_for_status()
        return r.content
    except Exception:
        return None

def _extract_pdf_text_from_bytes(bin_: bytes) -> str:
    """Tenta extrair texto do PDF primeiro com pdfminer; se falhar, usa PyPDF2."""
    try:
        # pdfminer é mais completo
        bio = io.BytesIO(bin_)
        txt = pdf_extract_text(bio)
        if txt and txt.strip():
            return txt.strip()
    except Exception:
        pass

    # fallback com PyPDF2
    try:
        bio2 = io.BytesIO(bin_)
        reader = PdfReader(bio2)
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                parts.append("")
        return "\n".join(parts).strip()
    except Exception:
        return ""

def _extract_pdf_text_from_url(u: str) -> str:
    bin_ = _fetch_bytes(u)
    if not bin_:
        return ""
    return _extract_pdf_text_from_bytes(bin_)

def _read_tabular_from_url(u: str, limit_rows: Optional[int] = None) -> Optional[pd.DataFrame]:
    """Mantido para compat: CSV/Sheets/Excel direto via URL (sem hiperlink por célula)."""
    try:
        if u.endswith(".xlsx") or u.endswith(".xls"):
            bin_ = requests.get(u, timeout=20).content
            df = pd.read_excel(io.BytesIO(bin_))
            if limit_rows:
                df = df.head(limit_rows)
            return df
        if "docs.google.com/spreadsheets" in u:
            if "export?format=csv" not in u:
                if "/edit" in u:
                    u = u.split("/edit")[0] + "/export?format=csv"
            bin_ = requests.get(u, timeout=20).content
            df = pd.read_csv(io.BytesIO(bin_))
            if limit_rows:
                df = df.head(limit_rows)
            return df
        if u.endswith(".csv"):
            bin_ = requests.get(u, timeout=20).content
            df = pd.read_csv(io.BytesIO(bin_))
            if limit_rows:
                df = df.head(limit_rows)
            return df
    except Exception:
        return None
    return None


# =========================
# Ingestão principal (com hiperlink -> PDF -> "Como fazer?")
# =========================
async def ingest_xlsx(file_bytes: bytes, usar_pdf_para_como_fazer: bool = True) -> Dict[str, Any]:
    """
    Lê planilha XLSX preservando hiperlinks e insere tarefas.
    Se usar_pdf_para_como_fazer=True, baixa o PDF do hiperlink de "Documento de Referência"
    e usa o texto extraído para preencher o campo "como_fazer" (quando este vier vazio).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active  # primeira aba

    # descobrir cabeçalho (primeira linha não vazia)
    header_row_idx = None
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row_vals):
            header_row_idx = r
            break
    if not header_row_idx:
        return {"criadas": 0, "erros": ["Não encontrei cabeçalho na planilha."]}

    headers: List[str] = []
    col_index_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row_idx, c).value
        name = (str(v).strip() if v is not None else f"COL_{c}")
        headers.append(name)
        col_index_map[name] = c

    def get_cell(row_idx: int, *names: str):
        for n in names:
            if n in col_index_map:
                return ws.cell(row_idx, col_index_map[n])
        return None

    def get_val(row_idx: int, *names: str):
        cell = get_cell(row_idx, *names)
        return None if cell is None else cell.value

    created, errors = 0, []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        # linha vazia? pula
        if not any(ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1)):
            continue
        try:
            nome_proj = get_val(r, "Nome do Projeto", "Projeto")
            if not nome_proj:
                errors.append(f"Linha {r}: 'Nome do Projeto' é obrigatório.")
                continue
            projeto = await Projeto.find_one(Projeto.nome == str(nome_proj).strip())
            if not projeto:
                errors.append(f"Linha {r}: Projeto '{nome_proj}' não encontrado.")
                continue

            email_resp = get_val(r, "Email Responsável", "Responsável", "Email")
            if not email_resp:
                errors.append(f"Linha {r}: 'Email Responsável' é obrigatório.")
                continue
            responsavel = await Funcionario.find_one(Funcionario.email == str(email_resp).strip())
            if not responsavel:
                errors.append(f"Linha {r}: Responsável '{email_resp}' não encontrado.")
                continue

            nome_tarefa = get_val(r, "Nome da Tarefa", "Tarefa", "Nome")
            if not nome_tarefa:
                errors.append(f"Linha {r}: 'Nome da Tarefa' é obrigatório.")
                continue

            # campos descritivos
            cell_como_fazer = get_cell(r, "Como fazer?", "Descrição")
            como_fazer = cell_como_fazer.value.strip() if (cell_como_fazer and isinstance(cell_como_fazer.value, str)) else None

            categoria = (get_val(r, "Categoria", "Classificação") or None)
            fase = (get_val(r, "Fase") or None)

            # documento de referência - CAPTURA DO HIPERLINK
            cell_doc = get_cell(r, "Documento de Referência", "Documento referência", "Documento")
            doc_ref_value = None if cell_doc is None else (cell_doc.value if not isinstance(cell_doc.value, str) else cell_doc.value.strip())
            doc_ref_link = None if cell_doc is None else (cell_doc.hyperlink.target if cell_doc.hyperlink else None)
            documento_referencia = doc_ref_link or doc_ref_value

            # prioridade (opcional)
            prioridade_raw = get_val(r, "Prioridade")
            prioridade = None
            if isinstance(prioridade_raw, str):
                p = prioridade_raw.strip().lower()
                if p == "media": p = "média"
                if p in {"baixa", "média", "alta"}:
                    prioridade = PrioridadeTarefa(p)

            # condição
            cond_raw = get_val(r, "Condição", "Condicao")
            cond = CondicaoTarefa.SEMPRE
            if isinstance(cond_raw, str):
                cc = cond_raw.strip().upper()
                if cc in {"SEMPRE", "A", "B", "C"}:
                    cond = CondicaoTarefa(cc)

            # porcentagem
            porc = get_val(r, "Porcentagem")
            if porc is None:
                porc = 100 if _normalize_bool(get_val(r, "Concluído", "Concluida")) else 0
            try:
                porc = int(float(porc))
            except Exception:
                porc = 0
            porc = max(0, min(100, porc))

            # datas
            dt_ini = get_val(r, "Data de Início")
            dt_fim = get_val(r, "Data de Fim")
            prazo_legado = get_val(r, "Prazo")
            exp_dias = get_val(r, "Exportação (dias)")

            dt_inicio = pd.to_datetime(dt_ini, errors="coerce") if dt_ini is not None else None
            dt_final = pd.to_datetime(dt_fim, errors="coerce") if dt_fim is not None else None

            if dt_final is pd.NaT or dt_final is None:
                prazo_dt = pd.to_datetime(prazo_legado, errors="coerce") if prazo_legado is not None else None
                if prazo_dt is not None and prazo_dt is not pd.NaT:
                    dt_final = prazo_dt

            if (dt_final is pd.NaT or dt_final is None) and dt_inicio is not None and pd.notna(dt_inicio) and exp_dias is not None:
                try:
                    dt_final = dt_inicio + pd.to_timedelta(int(exp_dias), unit="D")
                except Exception:
                    pass

            if dt_final is pd.NaT or dt_final is None:
                errors.append(f"Linha {r}: informe 'Data de Fim' ou 'Prazo' (legado) ou 'Data de Início' + 'Exportação (dias)'.")
                continue

            data_inicio_val = None if (dt_inicio is None or dt_inicio is pd.NaT) else dt_inicio.date()
            data_fim_val = dt_final.date()

            # ======== NOVO: usar o PDF do hiperlink como "Como fazer?" ========
            if not (como_fazer and como_fazer.strip()) and usar_pdf_para_como_fazer and documento_referencia and str(documento_referencia).lower().endswith(".pdf"):
                try:
                    pdf_texto = _extract_pdf_text_from_url(str(documento_referencia))
                    # Limpa o texto e reduz tamanho se necessário
                    como_fazer = (pdf_texto or "").strip()
                    if len(como_fazer) > 12000:
                        como_fazer = como_fazer[:12000] + "\n\n[...]"
                except Exception:
                    # mantém None se falhar
                    pass

            payload = TarefaCreate(
                nome=str(nome_tarefa).strip(),
                projeto_id=projeto.id,
                responsavel_id=responsavel.id,
                como_fazer=como_fazer,
                prioridade=prioridade,
                condicao=cond,
                categoria=categoria,
                porcentagem=porc,
                data_inicio=data_inicio_val,
                data_fim=data_fim_val,
                documento_referencia=documento_referencia,
                fase=fase,
                status=StatusTarefa.NAO_INICIADA if porc == 0 else (StatusTarefa.CONCLUIDA if porc == 100 else StatusTarefa.EM_ANDAMENTO),
            )

            tarefa = Tarefa(
                nome=payload.nome,
                projeto=projeto,
                responsavel=responsavel,
                como_fazer=payload.como_fazer,
                prioridade=payload.prioridade,
                condicao=payload.condicao or CondicaoTarefa.SEMPRE,
                categoria=payload.categoria,
                porcentagem=payload.porcentagem or 0,
                data_inicio=payload.data_inicio,
                data_fim=payload.data_fim or payload.prazo,
                documento_referencia=payload.documento_referencia,
                fase=payload.fase,
                status=payload.status or StatusTarefa.NAO_INICIADA,
            )
            if tarefa.status == StatusTarefa.CONCLUIDA:
                from datetime import datetime as _dt
                tarefa.dataConclusao = _dt.utcnow()

            await tarefa.insert()
            created += 1

        except Exception as e:
            errors.append(f"Linha {r}: {e}")

    return {"criadas": created, "erros": errors}


async def ingest_from_url(url: str) -> Dict[str, Any]:
    # Se for planilha: ingere (sem preservar hiperlink por célula)
    df = _read_tabular_from_url(url)
    if df is not None:
        # caminho antigo (sem hiperlink de célula); reusa dataframe
        return await _ingest_tasks_df_from_dataframe(df)

    # .docx
    if url.lower().endswith(".docx"):
        bin_ = _fetch_bytes(url)
        if not bin_:
            return {"tipo": "docx", "erro": "Falha ao baixar"}
        doc = DocxDocument(io.BytesIO(bin_))
        full_text = "\n".join(par.text for par in doc.paragraphs)
        return {"tipo": "docx", "bytes": len(bin_), "extract": full_text[:8000]}

    # HTML
    txt = _fetch_url_text(url)
    return {"tipo": "html", "extract": txt[:8000]}


# Caminho legado (quando já recebemos um DataFrame) — mantido para compat
async def _ingest_tasks_df_from_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    created, errors = 0, []
    df = df.rename(columns=lambda c: str(c).strip())

    def get(row, *names):
        for n in names:
            if n in row and pd.notna(row[n]):
                return row[n]
        return None

    for i, row in df.iterrows():
        try:
            nome_proj = get(row, "Nome do Projeto", "Projeto")
            if not nome_proj:
                errors.append(f"Linha {i+2}: 'Nome do Projeto' é obrigatório.")
                continue
            projeto = await Projeto.find_one(Projeto.nome == str(nome_proj).strip())
            if not projeto:
                errors.append(f"Linha {i+2}: Projeto '{nome_proj}' não encontrado.")
                continue

            email_resp = get(row, "Email Responsável", "Responsável", "Email")
            if not email_resp:
                errors.append(f"Linha {i+2}: 'Email Responsável' é obrigatório.")
                continue
            responsavel = await Funcionario.find_one(Funcionario.email == str(email_resp).strip())
            if not responsavel:
                errors.append(f"Linha {i+2}: Responsável '{email_resp}' não encontrado.")
                continue

            nome_tarefa = get(row, "Nome da Tarefa", "Tarefa", "Nome")
            if not nome_tarefa:
                errors.append(f"Linha {i+2}: 'Nome da Tarefa' é obrigatório.")
                continue

            como_fazer = get(row, "Como fazer?", "Descrição")
            categoria = get(row, "Categoria", "Classificação")
            fase = get(row, "Fase")
            documento_referencia = get(row, "Documento de Referência", "Documento referência", "Documento")

            prioridade_raw = get(row, "Prioridade")
            prioridade = None
            if isinstance(prioridade_raw, str):
                p = prioridade_raw.strip().lower()
                if p == "media": p = "média"
                if p in {"baixa", "média", "alta"}:
                    prioridade = PrioridadeTarefa(p)

            cond_raw = get(row, "Condição", "Condicao")
            cond = CondicaoTarefa.SEMPRE
            if isinstance(cond_raw, str):
                cc = cond_raw.strip().upper()
                if cc in {"SEMPRE", "A", "B", "C"}:
                    cond = CondicaoTarefa(cc)

            porc = get(row, "Porcentagem")
            if porc is None:
                porc = 100 if _normalize_bool(get(row, "Concluído", "Concluida")) else 0
            try:
                porc = int(float(porc))
            except Exception:
                porc = 0
            porc = max(0, min(100, porc))

            dt_ini = get(row, "Data de Início")
            dt_fim = get(row, "Data de Fim")
            prazo_legado = get(row, "Prazo")
            exp_dias = get(row, "Exportação (dias)")

            dt_inicio = pd.to_datetime(dt_ini, errors="coerce") if dt_ini is not None else None
            dt_final = pd.to_datetime(dt_fim, errors="coerce") if dt_fim is not None else None

            if dt_final is pd.NaT or dt_final is None:
                prazo_dt = pd.to_datetime(prazo_legado, errors="coerce") if prazo_legado is not None else None
                if prazo_dt is not None and prazo_dt is not pd.NaT:
                    dt_final = prazo_dt

            if (dt_final is pd.NaT or dt_final is None) and dt_inicio is not None and pd.notna(dt_inicio) and pd.notna(exp_dias):
                try:
                    dt_final = dt_inicio + pd.to_timedelta(int(exp_dias), unit="D")
                except Exception:
                    pass

            if dt_final is pd.NaT or dt_final is None:
                errors.append(f"Linha {i+2}: informe 'Data de Fim' ou 'Prazo' (legado) ou 'Data de Início' + 'Exportação (dias)'.")
                continue

            data_inicio_val = None if (dt_inicio is None or dt_inicio is pd.NaT) else dt_inicio.date()
            data_fim_val = dt_final.date()

            payload = TarefaCreate(
                nome=str(nome_tarefa).strip(),
                projeto_id=projeto.id,
                responsavel_id=responsavel.id,
                como_fazer=como_fazer,
                prioridade=prioridade,
                condicao=cond,
                categoria=categoria,
                porcentagem=porc,
                data_inicio=data_inicio_val,
                data_fim=data_fim_val,
                documento_referencia=documento_referencia,
                fase=fase,
                status=StatusTarefa.NAO_INICIADA if porc == 0 else (StatusTarefa.CONCLUIDA if porc == 100 else StatusTarefa.EM_ANDAMENTO),
            )

            tarefa = Tarefa(
                nome=payload.nome,
                projeto=projeto,
                responsavel=responsavel,
                como_fazer=payload.como_fazer,
                prioridade=payload.prioridade,
                condicao=payload.condicao or CondicaoTarefa.SEMPRE,
                categoria=payload.categoria,
                porcentagem=payload.porcentagem or 0,
                data_inicio=payload.data_inicio,
                data_fim=payload.data_fim or payload.prazo,
                documento_referencia=payload.documento_referencia,
                fase=payload.fase,
                status=payload.status or StatusTarefa.NAO_INICIADA,
            )
            if tarefa.status == StatusTarefa.CONCLUIDA:
                from datetime import datetime as _dt
                tarefa.dataConclusao = _dt.utcnow()

            await tarefa.insert()
            created += 1

        except Exception as e:
            errors.append(f"Linha {i+2}: {e}")

    return {"criadas": created, "erros": errors}
